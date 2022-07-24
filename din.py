from sys import argv
import json
import openpyxl
import requests
from bs4 import BeautifulSoup

def download_webpage(cin):
    try:
        URL = 'https://www.quickcompany.in/company/autocomplete?term=' + cin
        response = requests.get(URL)
        
        if response.status_code != 200:
            raise Exception('Cannot fetch webpage!')

        data = response.json()
        director = json.dumps(response.json()[0]['directors'][0]['slug'])
        return director[11:19]
        
        #for x in range (11,19):
        	#print(director[x], end='') #print only din horizontal
        	#out = (director[x], end='')
        	#return out
        	
        #print()
        #print(data[0]['directors'][0]['slug']) #print slug info
        #print()
        #fetched = json.dumps(response.json(), indent = 4, sort_keys =False) #prettyprint
        #print(fetched)
        #print()
        
        #fetched_dir = fetched["directors"]#get list of directors
        #for fetch in fetched_dir:
        #for x in range (433,450):
        #    print(fetched[x])
       
        
            
#         webpage = response.content
#         extract_content(webpage)

    except Exception as e:
        print(f'Error: {str(e)}')
        
path = "Book1.xlsx"
wb = openpyxl.load_workbook(path)
sheet = wb.active
max_r = sheet.max_row
print(max_r)
print()
for i in range(1, max_r + 1):
	cell = sheet.cell(row = i, column = 1)
	cin = cell.value
	din = download_webpage(cin)
	din_cell = sheet.cell(row = i, column = 2)
	din_cell.value = din
	
wb.save("Book1.xlsx")
print("Operation completed")

